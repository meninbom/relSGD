import time
import os
import glob
import subprocess
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import json
import sys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Função para formatar a data no formato dd/mm/aa
def formatar_data(data_str):
    return f"{data_str[:2]}/{data_str[2:4]}/{data_str[4:]}"

# Função para validar a data no formato ddmmaa
def validar_data(data_str):
    if len(data_str) != 6 or not data_str.isdigit():
        return False
    try:
        dia, mes, ano = int(data_str[:2]), int(data_str[2:4]), int(data_str[4:])
        datetime(2000 + ano, mes, dia)  # Assume século 21
        return True
    except ValueError:
        return False

# Função para carregar credenciais salvas
def carregar_credenciais():
    try:
        with open("config.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        return {"usuario": "", "senha": ""}

# Função para salvar credenciais
def salvar_credenciais(usuario, senha):
    with open("config.json", "w") as f:
        json.dump({"usuario": usuario, "senha": senha}, f)

# Função para encontrar o último arquivo baixado, excluindo arquivos existentes
def encontrar_ultimo_arquivo(diretorio, extensao="xlsx", arquivos_existentes=None):
    lista_arquivos = glob.glob(os.path.join(diretorio, f"*.{extensao}"))
    if arquivos_existentes:
        lista_arquivos = [f for f in lista_arquivos if f not in arquivos_existentes]
    if not lista_arquivos:
        return None
    return max(lista_arquivos, key=os.path.getctime)

# Função para aguardar o download do arquivo
def aguardar_download(diretorio, extensao="xlsx", timeout=90):
    arquivos_existentes = glob.glob(os.path.join(diretorio, f"*.{extensao}"))
    print(f"Arquivos existentes antes do download: {arquivos_existentes}")
    
    start_time = time.time()
    while time.time() - start_time < timeout:
        arquivo = encontrar_ultimo_arquivo(diretorio, extensao, arquivos_existentes)
        if arquivo and not arquivo.endswith(".crdownload"):
            print(f"Arquivo baixado detectado: {arquivo}")
            return arquivo
        time.sleep(1)
    print("Timeout atingido: nenhum arquivo baixado detectado.")
    return None

# Função para desbloquear o arquivo usando PowerShell
def desbloquear_arquivo(caminho_arquivo):
    try:
        comando = ["powershell", "-Command", "Unblock-File", "-Path", caminho_arquivo]
        subprocess.run(comando, check=True)
        print(f"Arquivo desbloqueado com sucesso: {caminho_arquivo}")
    except Exception as e:
        print(f"Erro ao desbloquear o arquivo: {e}")

# Função para renomear o arquivo, sempre gerando um novo arquivo com sufixo incremental
def renomear_arquivo(arquivo_atual, novo_nome_base):
    nome_base, extensao = os.path.splitext(novo_nome_base)
    diretorio = os.path.dirname(novo_nome_base)
    novo_nome = os.path.join(diretorio, f"{os.path.basename(nome_base)}{extensao}")
    contador = 1
    
    while os.path.exists(novo_nome):
        novo_nome = os.path.join(diretorio, f"{os.path.basename(nome_base)}_{contador}{extensao}")
        contador += 1
    
    try:
        os.rename(arquivo_atual, novo_nome)
        print(f"Arquivo renomeado para: {novo_nome}")
        return novo_nome
    except Exception as e:
        print(f"Erro ao renomear o arquivo: {e}")
        raise e

# Função para gerar gráfico de barras
def gerar_grafico_barras(tabela, caminho_arquivo):
    fig, ax = plt.subplots(figsize=(10, 6))
    tabela.plot(kind='bar', stacked=True, color=['green', 'red', 'orange'], ax=ax)
    plt.title("Distribuição de Situações por Técnico")
    plt.ylabel("Quantidade")
    plt.xlabel("Técnico")
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    grafico_path = caminho_arquivo.replace(".xlsx", "_grafico_barras.png")
    plt.savefig(grafico_path)
    plt.close()
    return grafico_path

# Função para gerar gráfico de pizza
# Função para gerar gráfico de pizza
def gerar_grafico_pizza(tabela, caminho_arquivo):
    soma_status = tabela.sum()
    fig, ax = plt.subplots()
    
    # Dicionário mapeando situações para cores
    mapeamento_cores = {
        'Concluído': '#008000',  # Verde
        'Prescrita': '#FF0000',  # Vermelho
        'Realizado': '#FFA500'   # Laranja
    }
    
    # Lista de cores na ordem dos índices de soma_status
    cores = [mapeamento_cores.get(situacao, '#CCCCCC') for situacao in soma_status.index]
    
    soma_status.plot(kind='pie', autopct='%1.1f%%', startangle=90, colors=cores, ax=ax)
    plt.title("Distribuição Geral por Situação")
    plt.ylabel("")
    plt.tight_layout()
    grafico_path = caminho_arquivo.replace(".xlsx", "_grafico_pizza.png")
    plt.savefig(grafico_path)
    plt.close()
    return grafico_path

# Função para gerar a tabela dinâmica e adicionar gráficos no Excel
def criar_tabela_dinamica(caminho_arquivo):
    desbloquear_arquivo(caminho_arquivo)

    df = pd.read_excel(caminho_arquivo)
    print(f"Colunas originais: {df.columns.tolist()}")

    responsavel_coluna = None
    situacao_coluna = None
    for coluna in df.columns:
        if 'responsável' in coluna.lower():
            responsavel_coluna = coluna
        if 'situação' in coluna.lower():
            situacao_coluna = coluna

    if responsavel_coluna and situacao_coluna:
        print(f"Colunas encontradas: Responsável -> {responsavel_coluna}, Situação -> {situacao_coluna}")
        
        responsaveis_excluir = [
            'Allan Cesar Gallan',
            'Allan Cesar Gallan ',
            'Amanda Paschoal Monteiro',
            'Nathan Wilian moreira',
            'Polyane Oliveira Nita',
            'Allan Cesar Galan ',
            'Tiago Rico Bocato',
            'Técnico DEVOPS (Implantação DW)',
            'Taynara Santos Viana',
            'Pricila Martins da Silva',
            'Gabriel dos Reis Conceição'
        ]
        
        df = df[~df[responsavel_coluna].isin(responsaveis_excluir)]
        print(f"Linhas após exclusão: {len(df)}")
        
        situacoes_validas = ['Realizado', 'Concluído', 'Prescrita']
        df_filtrado = df[df[situacao_coluna].isin(situacoes_validas)]

        tabela = pd.pivot_table(
            df_filtrado,
            index=responsavel_coluna,
            columns=situacao_coluna,
            aggfunc='size',
            fill_value=0
        )

        with pd.ExcelWriter(caminho_arquivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            tabela.to_excel(writer, sheet_name="DinamicTable")
        print("Tabela dinâmica adicionada com sucesso na planilha 'DinamicTable'.")

        grafico_barras_path = gerar_grafico_barras(tabela, caminho_arquivo)
        grafico_pizza_path = gerar_grafico_pizza(tabela, caminho_arquivo)

        wb = load_workbook(caminho_arquivo)
        ws = wb.create_sheet("Gráficos")

        ws.sheet_view.showGridLines = False

        img_barras = ExcelImage(grafico_barras_path)
        img_barras.width = 640
        img_barras.height = 480
        ws.add_image(img_barras, 'A1')

        img_pizza = ExcelImage(grafico_pizza_path)
        img_pizza.width = 480
        img_pizza.height = 400
        ws.add_image(img_pizza, 'L1')

        wb.save(caminho_arquivo)
        print("Gráficos adicionados na planilha 'Gráficos' com configurações aplicadas.")

        os.remove(grafico_barras_path)
        os.remove(grafico_pizza_path)
        print("Arquivos temporários de gráficos removidos.")
    else:
        print("Não foi possível identificar as colunas 'Responsável' e 'Situação'.")

# Função para executar o processo principal
def executar_script(usuario, senha, data_inicio_raw, data_fim_raw, navegador, download_dir):
    download_dir = os.path.normpath(download_dir)
    
    if not os.path.exists(download_dir):
        try:
            os.makedirs(download_dir, exist_ok=True)
        except Exception as e:
            raise PermissionError(f"Não foi possível criar o diretório: {download_dir}. Erro: {e}")
    if not os.access(download_dir, os.W_OK):
        raise PermissionError(f"Sem permissão de escrita no diretório: {download_dir}")
    
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-extensions")
    options.add_argument("--incognito")
    options.add_argument("--disable-logging")
    options.add_argument("--log-level=3")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-features=DownloadBubble,DownloadBubbleV2")
    
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "safebrowsing.disable_download_protection": True,
        "profile.default_content_settings.popups": 0,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_setting_values.notifications": 2,
        "profile.content_settings.exceptions.automatic_downloads.*.setting": 1,
        "download.restrictions": 0,
    }
    options.add_experimental_option("prefs", prefs)
    
    if navegador == "Brave":
        brave_path = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
        if not os.path.exists(brave_path):
            raise FileNotFoundError("Executável do Brave não encontrado no caminho especificado.")
        options.binary_location = brave_path
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    try:
        print(f"Iniciando processo com navegador: {navegador}")
        driver.get("https://sgd.dominiosistemas.com.br/login.html")
        user_input = driver.find_element(By.XPATH, '/html/body/div/form/input[2]')
        pwd_input = driver.find_element(By.XPATH, '/html/body/div/form/input[3]')
        login_btn = driver.find_element(By.XPATH, '/html/body/div/form/input[4]')
        user_input.send_keys(usuario)
        pwd_input.send_keys(senha)
        login_btn.click()
        
        time.sleep(3)
        driver.get("https://sgd.dominiosistemas.com.br/sgsc/faces/programacoes.html")
        
        data_inicio = formatar_data(data_inicio_raw)
        data_fim = formatar_data(data_fim_raw)
        
        data_inicio_input = driver.find_element(By.XPATH, '//*[@id="programacoesForm:agendadoDe"]')
        data_inicio_input.clear()
        data_inicio_input.click()
        data_inicio_input.send_keys(data_inicio)
        
        data_fim_input = driver.find_element(By.XPATH, '//*[@id="programacoesForm:agendadoAte"]')
        data_fim_input.clear()
        data_fim_input.click()
        data_fim_input.send_keys(data_fim)
        
        salvar_btn = driver.find_element(By.XPATH, '//*[@id="programacoesForm:atualizarBtn"]')
        salvar_btn.click()
        print("Datas salvas.")
        time.sleep(5)
        
        gerarelatorio_btn = driver.find_element(By.XPATH, '//*[@id="programacoesForm:gerarRelatorio"]')
        gerarelatorio_btn.click()
        time.sleep(5)
        
        # Aguardar o link de download estar visível
        print("Aguardando o link de download...")
        download_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="downloadBoxID"]/table/tbody/tr[5]/td/a'))
        )
        
        # Tentar download direto via URL
        download_url = download_link.get_attribute("href")
        print(f"URL de download obtida: {download_url}")
        driver.execute_script(f"window.location.href = '{download_url}';")
        
        print(f"Aguardando o download do arquivo em: {download_dir}")
        arquivo_baixado = aguardar_download(diretorio=download_dir, extensao="xlsx", timeout=90)
        if not arquivo_baixado:
            raise TimeoutError("O download do arquivo não foi concluído dentro do tempo limite.")
        
        driver.quit()
        
        if arquivo_baixado:
            nome_base = f"{data_inicio_raw}_{data_fim_raw}"
            base = os.path.join(download_dir, f"{nome_base}.xlsx")
            novo = renomear_arquivo(arquivo_baixado, base)
            print(f"Arquivo renomeado para: {novo}")
            criar_tabela_dinamica(novo)
        else:
            print("Nenhum arquivo .xlsx encontrado.")
    except Exception as e:
        print(f"Erro durante a execução: {e}")
        driver.quit()
        raise e

# Interface gráfica com tkinter
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Automação de Relatórios")
        self.root.geometry("510x280")

        style = ttk.Style()
        style.configure("TLabel", font=("Arial", 12))
        style.configure("TButton", font=("Arial", 12))

        credenciais = carregar_credenciais()
        usuario_inicial = credenciais["usuario"]
        senha_inicial = credenciais["senha"]

        ttk.Label(root, text="Usuário:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
        self.entry_usuario = ttk.Entry(root, width=30)
        self.entry_usuario.insert(0, usuario_inicial)
        self.entry_usuario.grid(row=0, column=1, padx=10, pady=5)

        ttk.Label(root, text="Senha:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
        self.entry_senha = ttk.Entry(root, width=30, show="*")
        self.entry_senha.insert(0, senha_inicial)
        self.entry_senha.grid(row=1, column=1, padx=10, pady=5)

        ttk.Label(root, text="Navegador:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
        self.combo_navegador = ttk.Combobox(root, values=["Brave", "Chrome"], width=27, state="readonly")
        self.combo_navegador.set("Brave")
        self.combo_navegador.grid(row=2, column=1, padx=10, pady=5)

        ttk.Label(root, text="Diretório de Download:").grid(row=3, column=0, padx=10, pady=5, sticky="e")
        self.entry_download_dir = ttk.Entry(root, width=30)
        self.entry_download_dir.insert(0, r"C:\Users\Dominio\Desktop\Python\executavel\downloads")
        self.entry_download_dir.grid(row=3, column=1, padx=10, pady=5)
        ttk.Button(root, text="Selecionar", command=self.selecionar_diretorio).grid(row=3, column=2, padx=5, pady=5)

        ttk.Label(root, text="Data Inicial (ddmmaa):").grid(row=4, column=0, padx=10, pady=5, sticky="e")
        self.entry_data_inicio = ttk.Entry(root, width=30)
        self.entry_data_inicio.grid(row=4, column=1, padx=10, pady=5)

        ttk.Label(root, text="Data Final (ddmmaa):").grid(row=5, column=0, padx=10, pady=5, sticky="e")
        self.entry_data_fim = ttk.Entry(root, width=30)
        self.entry_data_fim.grid(row=5, column=1, padx=10, pady=5)

        self.salvar_credenciais = tk.BooleanVar(value=True)
        ttk.Checkbutton(root, text="Salvar credenciais", variable=self.salvar_credenciais).grid(row=6, column=0, columnspan=2, pady=5)

        self.btn_executar = ttk.Button(root, text="Executar", command=self.executar)
        self.btn_executar.grid(row=7, column=0, columnspan=2, pady=10)

    def selecionar_diretorio(self):
        diretorio = filedialog.askdirectory(title="Selecione o Diretório de Download")
        if diretorio:
            self.entry_download_dir.delete(0, tk.END)
            self.entry_download_dir.insert(0, diretorio)

    def executar(self):
        usuario = self.entry_usuario.get().strip()
        senha = self.entry_senha.get().strip()
        navegador = self.combo_navegador.get()
        download_dir = self.entry_download_dir.get().strip()
        data_inicio = self.entry_data_inicio.get().strip()
        data_fim = self.entry_data_fim.get().strip()

        if not usuario or not senha:
            messagebox.showerror("Erro", "Usuário e senha são obrigatórios!")
            return
        if not download_dir:
            messagebox.showerror("Erro", "Selecione um diretório de download!")
            return
        if not validar_data(data_inicio):
            messagebox.showerror("Erro", "Data inicial inválida! Use o formato ddmmaa (ex.: 010123).")
            return
        if not validar_data(data_fim):
            messagebox.showerror("Erro", "Data final inválida! Use o formato ddmmaa (ex.: 311223).")
            return

        self.btn_executar.config(state="disabled")
        self.root.update()

        try:
            executar_script(usuario, senha, data_inicio, data_fim, navegador, download_dir)
            if self.salvar_credenciais.get():
                salvar_credenciais(usuario, senha)
            messagebox.showinfo("Sucesso", "Relatório Concluído")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha na execução: {str(e)}")
        finally:
            self.btn_executar.config(state="normal")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        caminho_arquivo = sys.argv[1]
        if os.path.exists(caminho_arquivo):
            criar_tabela_dinamica(caminho_arquivo)
        else:
            print(f"Erro: O arquivo '{caminho_arquivo}' não existe.")
    else:
        root = tk.Tk()
        app = App(root)
        root.mainloop()
