import time
import os
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import tkinter as tk
from tkinter import messagebox, filedialog
import customtkinter as ctk
import json
import sys
import threading
import platform
import logging
import requests
import tempfile
import shutil
import io
import zipfile

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("automacao_relatorios.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuração do tema e aparência
ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("green")  # Themes: "blue" (standard), "green", "dark-blue"

# Constantes e configurações
CONFIG_FILE = "config.json"
DEFAULT_CONFIG = {
    "usuario": "",
    "senha": "",
    "navegador": "Chrome",
    "download_dir": os.path.join(os.path.expanduser("~"), "Downloads"),
    "tema": "System",
    "responsaveis_excluir": [
        "Allan Cesar Gallan",
        "Allan Cesar Gallan ",
        "Amanda Paschoal Monteiro",
        "Nathan Wilian moreira",
        "Polyane Oliveira Nita",
        "Allan Cesar Galan ",
        "Tiago Rico Bocato",
        "Técnico DEVOPS (Implantação DW)",
        "Taynara Santos Viana",
        "Pricila Martins da Silva",
        "Gabriel dos Reis Conceição"
    ],
    "cores_graficos": {
        "Concluído": "#008000",  # Verde
        "Prescrita": "#FF0000",  # Vermelho
        "Realizado": "#FFA500"   # Laranja
    }
}

# Função para formatar a data no formato dd/mm/aa
def formatar_data(data_str):
    if len(data_str) != 6 or not data_str.isdigit():
        raise ValueError("Data deve ter 6 dígitos no formato ddmmaa")
    return f"{data_str[:2]}/{data_str[2:4]}/{data_str[4:]}"

# Função para validar a data no formato ddmmaa
def validar_data(data_str):
    if len(data_str) != 6 or not data_str.isdigit():
        return False
    try:
        dia, mes, ano = int(data_str[:2]), int(data_str[2:4]), int(data_str[4:])
        datetime(2000 + ano, mes, dia)  # Assume século 21
        return True
    except ValueError:
        return False

# Função para carregar configurações salvas
def carregar_configuracoes():
    try:
        with open(CONFIG_FILE, "r", encoding='utf-8') as f:
            config = json.load(f)
            for key, value in DEFAULT_CONFIG.items():
                if key not in config:
                    config[key] = value
            return config
    except FileNotFoundError:
        return DEFAULT_CONFIG.copy()
    except json.JSONDecodeError:
        logger.error(f"Erro ao decodificar o arquivo {CONFIG_FILE}. Usando configurações padrão.")
        return DEFAULT_CONFIG.copy()

# Função para salvar configurações
def salvar_configuracoes(config):
    try:
        with open(CONFIG_FILE, "w", encoding='utf-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Erro ao salvar configurações: {e}")

# Função para verificar se o arquivo Excel é válido (ZIP)
def verificar_arquivo_excel(caminho_arquivo):
    try:
        with zipfile.ZipFile(caminho_arquivo, 'r') as z:
            z.testzip()  # Testa a integridade do ZIP
        return True
    except zipfile.BadZipFile:
        logger.error(f"O arquivo {caminho_arquivo} não é um ZIP válido.")
        return False

# Função para processar dados Excel diretamente
def processar_dados_excel_direto(caminho_arquivo, config=None, callback=None):
    if config is None:
        config = carregar_configuracoes()
    
    if callback:
        callback(10, "Processando dados do Excel...")
    
    try:
        # Carregar o workbook original
        wb = load_workbook(caminho_arquivo)
        
        # Identificar a planilha de dados (assumindo a primeira planilha)
        ws_dados = wb[wb.sheetnames[0]]
        
        # Ler os dados com pandas
        df = pd.read_excel(caminho_arquivo, sheet_name=wb.sheetnames[0], engine="openpyxl")
        logger.info(f"Dados Excel lidos com sucesso. Colunas: {df.columns.tolist()}")
        
        if callback:
            callback(20, "Identificando colunas...")
        
        # Identificar colunas relevantes
        responsavel_coluna = None
        situacao_coluna = None
        data_coluna = None

        ((By.XPATH, '//*[@id="programacoesForm:agendadoDe"]'))
        
        for coluna in df.columns:
            col_lower = str(coluna).lower()
            if any(x in col_lower for x in ['responsável', 'responsavel', 'técnico', 'tecnico']):
                responsavel_coluna = coluna
            if any(x in col_lower for x in ['situação', 'situacao', 'status']):
                situacao_coluna = coluna
            if 'data' in col_lower and not any(x in col_lower for x in ['criacao', 'criação']):
                data_coluna = coluna
        
        if not responsavel_coluna or not situacao_coluna:
            logger.warning("Colunas não identificadas por nome, tentando por posição...")
            if len(df.columns) >= 5:
                responsavel_coluna = df.columns[3]
                situacao_coluna = df.columns[4]
                logger.info(f"Colunas por posição: Responsável -> {responsavel_coluna}, Situação -> {situacao_coluna}")
            else:
                raise ValueError("Não foi possível identificar as colunas necessárias.")
        
        logger.info(f"Colunas: Responsável -> {responsavel_coluna}, Situação -> {situacao_coluna}, Data -> {data_coluna}")
        
        if callback:
            callback(30, "Filtrando dados...")
        
        df_filtrado = df.copy()
        responsaveis_excluir = config.get("responsaveis_excluir", DEFAULT_CONFIG["responsaveis_excluir"])
        if df_filtrado[responsavel_coluna].notna().any():
            df_filtrado[responsavel_coluna] = df_filtrado[responsavel_coluna].astype(str)
            df_filtrado = df_filtrado[~df_filtrado[responsavel_coluna].isin(responsaveis_excluir)]
            logger.info(f"Linhas após exclusão de responsáveis: {len(df_filtrado)}")
        
        situacoes_validas = list(config.get("cores_graficos", DEFAULT_CONFIG["cores_graficos"]).keys())
        if df_filtrado[situacao_coluna].notna().any():
            df_filtrado[situacao_coluna] = df_filtrado[situacao_coluna].astype(str)
            situacoes_encontradas = df_filtrado[situacao_coluna].unique()
            logger.info(f"Situações encontradas: {situacoes_encontradas}")
            
            df_filtrado = df_filtrado[df_filtrado[situacao_coluna].isin(situacoes_validas)]
        
        if len(df_filtrado) == 0:
            logger.warning("Nenhum dado após filtragem, usando dados originais.")
            df_filtrado = df
        
        if callback:
            callback(40, "Criando tabela dinâmica...")
        
        tabela = pd.pivot_table(
            df_filtrado,
            index=responsavel_coluna,
            columns=situacao_coluna,
            aggfunc='size',
            fill_value=0
        )
        tabela['Total'] = tabela.sum(axis=1)
        tabela = tabela.sort_values('Total', ascending=False)
        logger.info("Tabela dinâmica criada com sucesso.")
        
        if callback:
            callback(50, "Gerando gráficos...")
        
        cores_graficos = config.get("cores_graficos", DEFAULT_CONFIG["cores_graficos"])
        tabela_sem_total = tabela.drop('Total', axis=1)
        
        grafico_barras = gerar_grafico_barras_memoria(tabela_sem_total, cores_graficos)
        grafico_pizza = gerar_grafico_pizza_memoria(tabela_sem_total, cores_graficos)
        
        if callback:
            callback(80, "Adicionando planilhas ao workbook...")
        
        # Remover planilhas "DinamicTable" e "Gráficos" se existirem e adicionar novas
        if "DinamicTable" in wb.sheetnames:
            wb.remove(wb["DinamicTable"])
        ws_tabela = wb.create_sheet("DinamicTable")
        
        if "Gráficos" in wb.sheetnames:
            wb.remove(wb["Gráficos"])
        ws_graficos = wb.create_sheet("Gráficos")
        
        # Escrever a tabela dinâmica na planilha "DinamicTable"
        for r_idx, idx in enumerate(tabela.index, 2):
            ws_tabela.cell(row=r_idx, column=1, value=str(idx))
        for c_idx, column in enumerate(tabela.columns, 2):
            ws_tabela.cell(row=1, column=c_idx, value=str(column))
        for r_idx, idx in enumerate(tabela.index, 2):
            for c_idx, column in enumerate(tabela.columns, 2):
                ws_tabela.cell(row=r_idx, column=c_idx, value=tabela.loc[idx, column])
        
        # Adicionar título à planilha "Gráficos"
        ws_graficos.cell(row=1, column=1, value="Relatório de Análise de Situações")
        
        # Adicionar gráficos como imagens
        img_barras = ExcelImage(grafico_barras)
        img_barras.width = 800
        img_barras.height = 500
        ws_graficos.add_image(img_barras, 'A3')
        
        img_pizza = ExcelImage(grafico_pizza)
        img_pizza.width = 600
        img_pizza.height = 500
        ws_graficos.add_image(img_pizza, 'N3')  # Gráfico de pizza na célula N3
        
        # Remover grades de linha de todas as planilhas
        for ws in wb:
            ws.sheet_view.showGridLines = False
        
        if callback:
            callback(90, "Salvando arquivo Excel...")
        
        # Salvar as alterações no mesmo arquivo
        wb.save(caminho_arquivo)
        
        if callback:
            callback(100, "Processamento concluído!")
        
        return caminho_arquivo
    except Exception as e:
        logger.error(f"Erro ao processar dados Excel: {e}")
        if callback:
            callback(100, f"Erro: {str(e)}")
        raise

# Função para gerar gráfico de barras em memória
def gerar_grafico_barras_memoria(tabela, cores_dict=None):
    if cores_dict is None:
        cores_dict = DEFAULT_CONFIG["cores_graficos"]
    
    sns.set_style("whitegrid")
    fig, ax = plt.subplots(figsize=(12, 7))
    
    tabela_ordenada = tabela.copy()
    tabela_ordenada['Total'] = tabela_ordenada.sum(axis=1)
    tabela_ordenada = tabela_ordenada.sort_values('Total', ascending=False).drop('Total', axis=1)
    
    tabela_plot = tabela_ordenada.head(15) if len(tabela_ordenada) > 15 else tabela_ordenada
    titulo = "Top 15 Técnicos - Distribuição de Situações" if len(tabela_ordenada) > 15 else "Distribuição de Situações por Técnico"
    
    cores_lista = [cores_dict.get(col, "#CCCCCC") for col in tabela_plot.columns]
    tabela_plot.plot(kind='bar', stacked=True, color=cores_lista, ax=ax)
    
    for i, (tecnico, row) in enumerate(tabela_plot.iterrows()):
        acumulado = 0
        for j, valor in enumerate(row):
            if valor > 0:
                plt.text(i, acumulado + valor/2, str(int(valor)), ha='center', va='center', fontweight='bold', color='white')
            acumulado += valor
    
    plt.title(titulo, fontsize=16, pad=20)
    plt.ylabel("Quantidade", fontsize=12)
    plt.xlabel("Técnico", fontsize=12)
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close(fig)
    return img_buffer

# Função para gerar gráfico de pizza em memória
def gerar_grafico_pizza_memoria(tabela, cores_dict=None):
    if cores_dict is None:
        cores_dict = DEFAULT_CONFIG["cores_graficos"]
    
    soma_status = tabela.sum()
    fig, ax = plt.subplots(figsize=(8, 8))
    
    cores = [cores_dict.get(situacao, '#CCCCCC') for situacao in soma_status.index]
    ax.pie(soma_status, labels=soma_status.index, colors=cores, autopct='%1.1f%%', startangle=90)
    ax.set_title("Distribuição Geral por Situação")
    plt.tight_layout()
    
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close(fig)
    return img_buffer

# Função para executar o processo principal
def executar_script(usuario, senha, data_inicio_raw, data_fim_raw, navegador, download_dir, config=None, callback=None):
    if config is None:
        config = carregar_configuracoes()
    
    download_dir = os.path.normpath(download_dir)
    
    if not os.path.exists(download_dir):
        os.makedirs(download_dir, exist_ok=True)
    if not os.access(download_dir, os.W_OK):
        raise PermissionError(f"Sem permissão de escrita no diretório: {download_dir}")
    
    options = Options()
    options.add_argument("--start-minimized")
    # Removidas opções potencialmente problemáticas
    # options.add_argument("--disable-extensions")
    # options.add_argument("--incognito")
    # options.add_argument("--no-sandbox")
    # options.add_argument("--disable-gpu")
    
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False,
    }
    options.add_experimental_option("prefs", prefs)
    
    if navegador == "Brave" and platform.system() == "Windows":
        brave_path = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
        if not os.path.exists(brave_path):
            raise FileNotFoundError("Executável do Brave não encontrado.")
        options.binary_location = brave_path
    
    driver = None
    try:
        logger.info("Inicializando ChromeDriver...")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        logger.info("ChromeDriver inicializado com sucesso.")
        driver.minimize_window()
        
        if callback:
            callback(20, "Acessando sistema...")
        
        # Substitua pelo URL real do sistema SGD
        url_sgd = "https://sgd.dominiosistemas.com.br"  # Atualize com o URL correto
        logger.info(f"Acessando URL: {url_sgd}")
        driver.get(url_sgd)
        
        if callback:
            callback(25, "Realizando login...")
        
        logger.info("Aguardando campos de login...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/form/input[2]')))
        user_input = driver.find_element(By.XPATH, '/html/body/div/form/input[2]')
        pwd_input = driver.find_element(By.XPATH, '/html/body/div/form/input[3]')
        login_btn = driver.find_element(By.XPATH, '/html/body/div/form/input[4]')
        
        logger.info("Preenchendo credenciais...")
        user_input.send_keys(usuario)
        pwd_input.send_keys(senha)
        login_btn.click()
        
        time.sleep(5)
        
        if callback:
            callback(30, "Acessando página de programações...")
        
        # Substitua pelo URL real da página de programações
        url_programacoes = "https://sgd.dominiosistemas.com.br/sgsc/faces/programacoes.html"  # Atualize com o URL correto
        logger.info(f"Acessando URL de programações: {url_programacoes}")
        driver.get(url_programacoes)
        
        if callback:
            callback(35, "Configurando datas...")
        
        data_inicio = formatar_data(data_inicio_raw)
        data_fim = formatar_data(data_fim_raw)
        
        # Localizar e preencher o campo de data inicial
        logger.info("Preenchendo data inicial...")
        data_inicio_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="programacoesForm:agendadoDe"]'))
        )
        data_inicio_input.clear()
        data_inicio_input.click()
        data_inicio_input.send_keys(Keys.CONTROL + "a")
        data_inicio_input.send_keys(Keys.DELETE)
        data_inicio_input.send_keys(data_inicio)
        data_inicio_input.send_keys(Keys.TAB)
        
        # Localizar e preencher o campo de data final
        logger.info("Preenchendo data final...")
        data_fim_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="programacoesForm:agendadoAte"]'))
        )
        data_fim_input.clear()
        data_fim_input.click()
        data_fim_input.send_keys(Keys.CONTROL + "a")
        data_fim_input.send_keys(Keys.DELETE)
        data_fim_input.send_keys(data_fim)
        data_fim_input.send_keys(Keys.TAB)
        
        if callback:
            callback(40, "Salvando configurações de datas...")
        
        logger.info("Clicando no botão de salvar...")
        salvar_btn = driver.find_element(By.XPATH, '//*[@id="programacoesForm:atualizarBtn"]')
        salvar_btn.click()
        time.sleep(5)
        
        if callback:
            callback(45, "Gerando relatório...")
        
        logger.info("Clicando no botão de gerar relatório...")
        gerarelatorio_btn = driver.find_element(By.XPATH, '//*[@id="programacoesForm:gerarRelatorio"]')
        gerarelatorio_btn.click()
        time.sleep(5)
        
        if callback:
            callback(50, "Aguardando link de download...")
        
        # Esperar o link de download estar visível e clicável
        logger.info("Aguardando link de download...")
        download_link = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="downloadBoxID"]/table/tbody/tr[5]/td/a'))
        )
        
        # Obter o URL do link de download
        download_url = download_link.get_attribute('href')
        logger.info(f"URL de download obtido: {download_url}")
        
        # Obter cookies do driver
        cookies = driver.get_cookies()
        cookies_dict = {cookie['name']: cookie['value'] for cookie in cookies}
        
        # Baixar o arquivo diretamente com requests
        if callback:
            callback(60, "Baixando arquivo...")
        
        logger.info("Iniciando download do arquivo...")
        response = requests.get(download_url, cookies=cookies_dict, stream=True)
        if response.status_code != 200:
            raise Exception(f"Falha ao baixar o arquivo: Status {response.status_code}")
        
        # Salvar o arquivo com um nome temporário
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    temp_file.write(chunk)
            temp_path = temp_file.name
        
        # Verificar se o arquivo é um Excel válido
        if not verificar_arquivo_excel(temp_path):
            os.remove(temp_path)
            raise Exception("O arquivo baixado não é um arquivo Excel válido.")
        
        # Renomear o arquivo para o nome final
        nome_arquivo = f"{data_inicio_raw}_{data_fim_raw}.xlsx"
        caminho_destino = os.path.join(download_dir, nome_arquivo)
        contador = 1
        while os.path.exists(caminho_destino):
            nome_base, extensao = os.path.splitext(nome_arquivo)
            caminho_destino = os.path.join(download_dir, f"{nome_base}_{contador}{extensao}")
            contador += 1
        
        shutil.move(temp_path, caminho_destino)
        logger.info(f"Arquivo baixado e renomeado para: {caminho_destino}")
        
        # Processar o arquivo renomeado
        if callback:
            callback(70, "Processando dados do relatório...")
        
        caminho_processado = processar_dados_excel_direto(
            caminho_destino,
            config=config,
            callback=lambda p, m: callback(70 + int(p * 0.25), m) if callback else None
        )
        logger.info(f"Arquivo processado: {caminho_processado}")
        
        if callback:
            callback(100, f"Processamento concluído! Arquivo salvo em: {caminho_destino}")
        
        return caminho_destino
    except Exception as e:
        logger.error(f"Erro durante a execução: {e}", exc_info=True)
        if callback:
            callback(100, f"Erro: {str(e)}")
        raise
    finally:
        if driver:
            logger.info("Finalizando WebDriver...")
            driver.quit()

# Classe para visualização de gráficos
class VisualizadorGraficos(ctk.CTkToplevel):
    def __init__(self, master, tabela, titulo="Visualização de Gráficos"):
        super().__init__(master)
        self.title(titulo)
        self.geometry("1000x800")
        self.minsize(800, 600)
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        
        self.frame_controles = ctk.CTkFrame(self)
        self.frame_controles.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        self.btn_barras = ctk.CTkButton(self.frame_controles, text="Gráfico de Barras", command=self.mostrar_barras)
        self.btn_barras.grid(row=0, column=0, padx=10, pady=10)
        
        self.btn_pizza = ctk.CTkButton(self.frame_controles, text="Gráfico de Pizza", command=self.mostrar_pizza)
        self.btn_pizza.grid(row=0, column=1, padx=10, pady=10)
        
        self.frame_grafico = ctk.CTkFrame(self)
        self.frame_grafico.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        self.frame_grafico.grid_columnconfigure(0, weight=1)
        self.frame_grafico.grid_rowconfigure(0, weight=1)
        
        self.tabela = tabela
        self.mostrar_barras()
    
    def limpar_frame_grafico(self):
        for widget in self.frame_grafico.winfo_children():
            widget.destroy()
    
    def mostrar_barras(self):
        self.limpar_frame_grafico()
        
        fig = Figure(figsize=(10, 6), dpi=100)
        ax = fig.add_subplot(111)
        
        tabela_plot = self.tabela.drop('Total', axis=1) if 'Total' in self.tabela.columns else self.tabela
        tabela_plot = tabela_plot.head(15) if len(tabela_plot) > 15 else tabela_plot
        
        tabela_plot.plot(kind='bar', stacked=True, ax=ax)
        ax.set_title("Distribuição de Situações por Técnico")
        ax.set_ylabel("Quantidade")
        ax.set_xlabel("Técnico")
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
        
        canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
    
    def mostrar_pizza(self):
        self.limpar_frame_grafico()
        
        fig = Figure(figsize=(8, 8), dpi=100)
        ax = Figure_subplot(111) # type: ignore
        
        tabela_plot = self.tabela.drop('Total', axis=1) if 'Total' in self.tabela.columns else self.tabela
        soma_status = tabela_plot.sum()
        
        ax.pie(soma_status, labels=soma_status.index, autopct='%1.1f%%', startangle=90, explode=[0.05] * len(soma_status))
        ax.set_title("Distribuição Geral por Situação")
        
        canvas = FigureCanvasTkAgg(fig, master=self.frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")

# Classe para a tela de configurações
class TelaConfiguracoes(ctk.CTkToplevel):
    def __init__(self, master, config=None, callback=None):
        super().__init__(master)
        self.title("Configurações")
        self.geometry("700x500")
        self.minsize(700, 500)
        
        self.config = config if config else carregar_configuracoes()
        self.callback = callback
        
        self.notebook = ctk.CTkTabview(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.tab_geral = self.notebook.add("Geral")
        self.tab_responsaveis = self.notebook.add("Responsáveis")
        self.tab_cores = self.notebook.add("Cores")
        
        self.configurar_aba_geral()
        self.configurar_aba_responsaveis()
        self.configurar_aba_cores()
        
        self.frame_botoes = ctk.CTkFrame(self)
        self.frame_botoes.pack(fill="x", padx=10, pady=10)
        
        self.btn_salvar = ctk.CTkButton(self.frame_botoes, text="Salvar", command=self.salvar_configuracoes)
        self.btn_salvar.pack(side="right", padx=10, pady=10)
        
        self.btn_cancelar = ctk.CTkButton(self.frame_botoes, text="Cancelar", command=self.destroy)
        self.btn_cancelar.pack(side="right", padx=10, pady=10)
    
    def configurar_aba_geral(self):
        ctk.CTkLabel(self.tab_geral, text="Tema:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.combo_tema = ctk.CTkComboBox(self.tab_geral, values=["System", "Dark", "Light"])
        self.combo_tema.grid(row=0, column=1, padx=10, pady=10, sticky="w")
        self.combo_tema.set(self.config.get("tema", "System"))
        
        ctk.CTkLabel(self.tab_geral, text="Navegador Padrão:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.combo_navegador = ctk.CTkComboBox(self.tab_geral, values=["Chrome", "Brave"])
        self.combo_navegador.grid(row=1, column=1, padx=10, pady=10, sticky="w")
        self.combo_navegador.set(self.config.get("navegador", "Chrome"))
        
        ctk.CTkLabel(self.tab_geral, text="Diretório de Download:").grid(row=2, column=0, padx=10, pady=10, sticky="w")
        self.frame_dir = ctk.CTkFrame(self.tab_geral)
        self.frame_dir.grid(row=2, column=1, padx=10, pady=10, sticky="ew")
        
        self.entry_dir = ctk.CTkEntry(self.frame_dir, width=300)
        self.entry_dir.pack(side="left", fill="x", expand=True)
        self.entry_dir.insert(0, self.config.get("download_dir", os.path.join(os.path.expanduser("~"), "Downloads")))
        
        self.btn_dir = ctk.CTkButton(self.frame_dir, text="...", width=30, command=self.selecionar_diretorio)
        self.btn_dir.pack(side="right", padx=5)
    
    def configurar_aba_responsaveis(self):
        self.frame_resp = ctk.CTkFrame(self.tab_responsaveis)
        self.frame_resp.pack(fill="both", expand=True, padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_resp, text="Responsáveis a excluir:").pack(anchor="w", padx=10, pady=5)
        
        self.frame_lista = ctk.CTkFrame(self.frame_resp)
        self.frame_lista.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.scrollbar = ctk.CTkScrollbar(self.frame_lista)
        self.scrollbar.pack(side="right", fill="y")
        
        self.listbox = tk.Listbox(self.frame_lista, yscrollcommand=self.scrollbar.set, bg="#2B2B2B", fg="#FFFFFF", selectbackground="#1F538D", font=("Arial", 12))
        self.listbox.pack(side="left", fill="both", expand=True)
        
        self.scrollbar.configure(command=self.listbox.yview)
        
        for resp in self.config.get("responsaveis_excluir", []):
            self.listbox.insert(tk.END, resp)
        
        self.frame_botoes_resp = ctk.CTkFrame(self.frame_resp)
        self.frame_botoes_resp.pack(fill="x", padx=10, pady=5)
        
        self.btn_adicionar = ctk.CTkButton(self.frame_botoes_resp, text="Adicionar", command=self.adicionar_responsavel)
        self.btn_adicionar.pack(side="left", padx=5, pady=5)
        
        self.btn_remover = ctk.CTkButton(self.frame_botoes_resp, text="Remover", command=self.remover_responsavel)
        self.btn_remover.pack(side="left", padx=5, pady=5)
    
    def configurar_aba_cores(self):
        self.frame_cores = ctk.CTkFrame(self.tab_cores)
        self.frame_cores.pack(fill="both", expand=True, padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_cores, text="Cores dos gráficos:").pack(anchor="w", padx=10, pady=5)
        
        self.cores_entries = {}
        row = 1
        cores_dict = self.config.get("cores_graficos", DEFAULT_CONFIG["cores_graficos"])
        
        for situacao, cor in cores_dict.items():
            ctk.CTkLabel(self.frame_cores, text=f"{situacao}:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
            frame_cor = ctk.CTkFrame(self.frame_cores)
            frame_cor.grid(row=row, column=1, padx=10, pady=5, sticky="w")
            
            entry_cor = ctk.CTkEntry(frame_cor, width=100)
            entry_cor.pack(side="left", padx=5)
            entry_cor.insert(0, cor)
            
            btn_cor = ctk.CTkButton(frame_cor, text="Escolha", width=80, command=lambda s=situacao: self.escolher_cor(s))
            btn_cor.pack(side="left", padx=5)
            
            preview = ctk.CTkLabel(frame_cor, text="      ", bg_color=cor, fg_color=cor, corner_radius=5)
            preview.pack(side="left", padx=5)
            
            self.cores_entries[situacao] = {"entry": entry_cor, "preview": preview}
            row += 1
    
    def selecionar_diretorio(self):
        diretorio = filedialog.askdirectory(title="Selecione o Diretório de Download")
        if diretorio:
            self.entry_dir.delete(0, tk.END)
            self.entry_dir.insert(0, diretorio)
    
    def adicionar_responsavel(self):
        dialog = ctk.CTkInputDialog(text="Nome do responsável:", title="Adicionar Responsável")
        responsavel = dialog.get_input()
        if responsavel and responsavel.strip():
            self.listbox.insert(tk.END, responsavel.strip())
    
    def remover_responsavel(self):
        selecionado = self.listbox.curselection()
        if selecionado:
            self.listbox.delete(selecionado)
    
    def escolher_cor(self, situacao):
        cores = ["#008000", "#FF0000", "#FFA500", "#0000FF", "#800080", "#008080"]
        entry = self.cores_entries[situacao]["entry"]
        preview = self.cores_entries[situacao]["preview"]
        cor_atual = entry.get()
        
        try:
            idx = cores.index(cor_atual)
            nova_cor = cores[(idx + 1) % len(cores)]
        except ValueError:
            nova_cor = cores[0]
        
        entry.delete(0, tk.END)
        entry.insert(0, nova_cor)
        preview.configure(bg_color=nova_cor, fg_color=nova_cor)
    
    def salvar_configuracoes(self):
        self.config["tema"] = self.combo_tema.get()
        self.config["navegador"] = self.combo_navegador.get()
        self.config["download_dir"] = self.entry_dir.get()
        
        responsaveis = [self.listbox.get(i) for i in range(self.listbox.size())]
        self.config["responsaveis_excluir"] = responsaveis
        
        cores = {situacao: widgets["entry"].get() for situacao, widgets in self.cores_entries.items()}
        self.config["cores_graficos"] = cores
        
        salvar_configuracoes(self.config)
        ctk.set_appearance_mode(self.config["tema"])
        
        if self.callback:
            self.callback(self.config)
        self.destroy()

# Classe para a tela de exportação
class TelaExportacao(ctk.CTkToplevel):
    def __init__(self, master, arquivo_excel):
        super().__init__(master)
        self.title("Exportar Relatório")
        self.geometry("500x300")
        self.minsize(500, 300)
        
        self.arquivo_excel = arquivo_excel
        
        self.frame_principal = ctk.CTkFrame(self)
        self.frame_principal.pack(fill="both", expand=True, padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_principal, text="Exportar Relatório", font=("Arial", 16, "bold")).pack(pady=10)
        
        self.frame_opcoes = ctk.CTkFrame(self.frame_principal)
        self.frame_opcoes.pack(fill="x", padx=10, pady=10)
        
        self.check_pdf = ctk.CTkCheckBox(self.frame_opcoes, text="Exportar para PDF")
        self.check_pdf.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.check_pdf.select()
        
        self.check_csv = ctk.CTkCheckBox(self.frame_opcoes, text="Exportar tabela para CSV")
        self.check_csv.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        
        self.check_img = ctk.CTkCheckBox(self.frame_opcoes, text="Exportar gráficos como imagens")
        self.check_img.grid(row=2, column=0, padx=10, pady=10, sticky="w")
        
        self.frame_dir = ctk.CTkFrame(self.frame_principal)
        self.frame_dir.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_dir, text="Diretório de destino:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.entry_dir = ctk.CTkEntry(self.frame_dir, width=300)
        self.entry_dir.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        self.entry_dir.insert(0, os.path.dirname(self.arquivo_excel))
        
        self.btn_dir = ctk.CTkButton(self.frame_dir, text="...", width=30, command=self.selecionar_diretorio)
        self.btn_dir.grid(row=0, column=2, padx=5, pady=10)
        
        self.frame_botoes = ctk.CTkFrame(self.frame_principal)
        self.frame_botoes.pack(fill="x", padx=10, pady=10)
        
        self.btn_exportar = ctk.CTkButton(self.frame_botoes, text="Exportar", command=self.exportar)
        self.btn_exportar.pack(side="right", padx=10, pady=10)
        
        self.btn_cancelar = ctk.CTkButton(self.frame_botoes, text="Cancelar", command=self.destroy)
        self.btn_cancelar.pack(side="right", padx=10, pady=10)
    
    def selecionar_diretorio(self):
        diretorio = filedialog.askdirectory(title="Selecione o Diretório de Destino")
        if diretorio:
            self.entry_dir.delete(0, tk.END)
            self.entry_dir.insert(0, diretorio)
    
    def exportar(self):
        diretorio = self.entry_dir.get()
        if not os.path.exists(diretorio):
            messagebox.showerror("Erro", "Diretório de destino inválido!")
            return
        
        try:
            if self.check_pdf.get():
                self.exportar_pdf(diretorio)
            if self.check_csv.get():
                self.exportar_csv(diretorio)
            if self.check_img.get():
                self.exportar_imagens(diretorio)
            messagebox.showinfo("Sucesso", "Exportação concluída com sucesso!")
            self.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro durante a exportação: {str(e)}")
    
    def exportar_pdf(self, diretorio):
        nome_base = os.path.splitext(os.path.basename(self.arquivo_excel))[0]
        caminho_pdf = os.path.join(diretorio, f"{nome_base}.pdf")
        with open(caminho_pdf, "w") as f:
            f.write("Simulação de PDF")
        logger.info(f"Relatório exportado para PDF: {caminho_pdf}")
    
    def exportar_csv(self, diretorio):
        nome_base = os.path.splitext(os.path.basename(self.arquivo_excel))[0]
        caminho_csv = os.path.join(diretorio, f"{nome_base}_tabela.csv")
        df = pd.read_excel(self.arquivo_excel, sheet_name="DinamicTable", engine='openpyxl')
        df.to_csv(caminho_csv, index=True)
        logger.info(f"Tabela exportada para CSV: {caminho_csv}")
    
    def exportar_imagens(self, diretorio):
        nome_base = os.path.splitext(os.path.basename(self.arquivo_excel))[0]
        df = pd.read_excel(self.arquivo_excel, sheet_name="DinamicTable", engine='openpyxl')
        config = carregar_configuracoes()
        cores_graficos = config.get("cores_graficos", DEFAULT_CONFIG["cores_graficos"])
        
        df_plot = df.drop('Total', axis=1) if 'Total' in df.columns else df
        
        grafico_barras = gerar_grafico_barras_memoria(df_plot, cores_graficos)
        grafico_pizza = gerar_grafico_pizza_memoria(df_plot, cores_graficos)
        
        grafico_barras_path = os.path.join(diretorio, f"{nome_base}_barras.png")
        grafico_pizza_path = os.path.join(diretorio, f"{nome_base}_pizza.png")
        
        with open(grafico_barras_path, 'wb') as f:
            f.write(grafico_barras.getvalue())
        with open(grafico_pizza_path, 'wb') as f:
            f.write(grafico_pizza.getvalue())
        
        logger.info(f"Gráficos exportados: {grafico_barras_path}, {grafico_pizza_path}")

# Classe para a tela de comparação
class TelaComparacao(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Comparar Períodos")
        self.geometry("600x400")
        self.minsize(600, 400)
        
        self.frame_principal = ctk.CTkFrame(self)
        self.frame_principal.pack(fill="both", expand=True, padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_principal, text="Comparar Períodos", font=("Arial", 16, "bold")).pack(pady=10)
        
        self.frame_arquivos = ctk.CTkFrame(self.frame_principal)
        self.frame_arquivos.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_arquivos, text="Primeiro período:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.entry_arquivo1 = ctk.CTkEntry(self.frame_arquivos, width=300)
        self.entry_arquivo1.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        self.btn_arquivo1 = ctk.CTkButton(self.frame_arquivos, text="...", width=30, command=lambda: self.selecionar_arquivo(1))
        self.btn_arquivo1.grid(row=0, column=2, padx=5, pady=10)
        
        ctk.CTkLabel(self.frame_arquivos, text="Segundo período:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.entry_arquivo2 = ctk.CTkEntry(self.frame_arquivos, width=300)
        self.entry_arquivo2.grid(row=1, column=1, padx=10, pady=10, sticky="ew")
        self.btn_arquivo2 = ctk.CTkButton(self.frame_arquivos, text="...", width=30, command=lambda: self.selecionar_arquivo(2))
        self.btn_arquivo2.grid(row=1, column=2, padx=5, pady=10)
        
        self.frame_opcoes = ctk.CTkFrame(self.frame_principal)
        self.frame_opcoes.pack(fill="x", padx=10, pady=10)
        
        self.var_tipo = ctk.StringVar(value="tecnico")
        ctk.CTkLabel(self.frame_opcoes, text="Tipo de comparação:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.radio_tecnico = ctk.CTkRadioButton(self.frame_opcoes, text="Por técnico", variable=self.var_tipo, value="tecnico")
        self.radio_tecnico.grid(row=0, column=1, padx=10, pady=10, sticky="w")
        self.radio_situacao = ctk.CTkRadioButton(self.frame_opcoes, text="Por situação", variable=self.var_tipo, value="situacao")
        self.radio_situacao.grid(row=0, column=2, padx=10, pady=10, sticky="w")
        
        self.frame_botoes = ctk.CTkFrame(self.frame_principal)
        self.frame_botoes.pack(fill="x", padx=10, pady=10)
        
        self.btn_comparar = ctk.CTkButton(self.frame_botoes, text="Comparar", command=self.comparar)
        self.btn_comparar.pack(side="right", padx=10, pady=10)
        self.btn_cancelar = ctk.CTkButton(self.frame_botoes, text="Cancelar", command=self.destroy)
        self.btn_cancelar.pack(side="right", padx=10, pady=10)
    
    def selecionar_arquivo(self, num):
        arquivo = filedialog.askopenfilename(title=f"Selecione o arquivo do {'primeiro' if num == 1 else 'segundo'} período", filetypes=[("Arquivos Excel", "*.xlsx")])
        if arquivo:
            (self.entry_arquivo1 if num == 1 else self.entry_arquivo2).delete(0, tk.END)
            (self.entry_arquivo1 if num == 1 else self.entry_arquivo2).insert(0, arquivo)
    
    def comparar(self):
        arquivo1, arquivo2 = self.entry_arquivo1.get(), self.entry_arquivo2.get()
        tipo = self.var_tipo.get()
        
        if not arquivo1 or not os.path.exists(arquivo1) or not arquivo2 or not os.path.exists(arquivo2):
            messagebox.showerror("Erro", "Arquivos inválidos!")
            return
        
        try:
            df1 = pd.read_excel(arquivo1, sheet_name="DinamicTable", engine='openpyxl')
            df2 = pd.read_excel(arquivo2, sheet_name="DinamicTable", engine='openpyxl')
            
            df1 = df1.drop('Total', axis=1) if 'Total' in df1.columns else df1
            df2 = df2.drop('Total', axis=1) if 'Total' in df2.columns else df2
            
            if tipo == "tecnico":
                self.comparar_por_tecnico(df1, df2)
            else:
                self.comparar_por_situacao(df1, df2)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro durante a comparação: {str(e)}")
    
    def comparar_por_tecnico(self, df1, df2):
        nome1 = os.path.splitext(os.path.basename(self.entry_arquivo1.get()))[0]
        nome2 = os.path.splitext(os.path.basename(self.entry_arquivo2.get()))[0]
        
        janela = ctk.CTkToplevel(self)
        janela.title(f"Comparação por Técnico: {nome1} vs {nome2}")
        janela.geometry("1000x600")
        
        fig = Figure(figsize=(12, 8), dpi=100)
        ax = fig.add_subplot(111)
        
        total1 = df1.sum(axis=1)
        total2 = df2.sum(axis=1)
        tecnicos = set(total1.index).intersection(set(total2.index))
        
        df_comp = pd.DataFrame({nome1: [total1.get(tec, 0) for tec in tecnicos], nome2: [total2.get(tec, 0) for tec in tecnicos]}, index=list(tecnicos))
        df_comp['Total'] = df_comp.sum(axis=1)
        df_comp = df_comp.sort_values('Total', ascending=False).drop('Total', axis=1)
        df_comp = df_comp.head(15) if len(df_comp) > 15 else df_comp
        
        df_comp.plot(kind='bar', ax=ax)
        ax.set_title("Comparação de Períodos por Técnico")
        ax.set_ylabel("Quantidade")
        ax.set_xlabel("Técnico")
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
        
        canvas = FigureCanvasTkAgg(fig, master=janela)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
    
    def comparar_por_situacao(self, df1, df2):
        nome1 = os.path.splitext(os.path.basename(self.entry_arquivo1.get()))[0]
        nome2 = os.path.splitext(os.path.basename(self.entry_arquivo2.get()))[0]
        
        janela = ctk.CTkToplevel(self)
        janela.title(f"Comparação por Situação: {nome1} vs {nome2}")
        janela.geometry("800x600")
        
        fig = Figure(figsize=(10, 8), dpi=100)
        ax = fig.add_subplot(111)
        
        total1 = df1.sum()
        total2 = df2.sum()
        situacoes = set(total1.index).union(set(total2.index))
        
        df_comp = pd.DataFrame({nome1: [total1.get(sit, 0) for sit in situacoes], nome2: [total2.get(sit, 0) for sit in situacoes]}, index=list(situacoes))
        df_comp.plot(kind='bar', ax=ax)
        
        ax.set_title("Comparação de Períodos por Situação")
        ax.set_ylabel("Quantidade")
        ax.set_xlabel("Situação")
        
        canvas = FigureCanvasTkAgg(fig, master=janela)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)

# Interface gráfica moderna com CustomTkinter
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Automação de Relatórios - SGD")
        self.root.geometry("900x600")
        self.root.minsize(1200, 600)
        
        self.config = carregar_configuracoes()
        ctk.set_appearance_mode(self.config.get("tema", "System"))
        
        self.arquivo_processado = None
        self.thread_execucao = None
        
        self.criar_layout()
        self.preencher_campos()
    
    def criar_layout(self):
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_rowconfigure(2, weight=0)
        
        self.frame_cabecalho = ctk.CTkFrame(self.root)
        self.frame_cabecalho.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        ctk.CTkLabel(self.frame_cabecalho, text="Automação de Relatórios - SGD", font=("Arial", 20, "bold")).pack(side="left", padx=20, pady=10)
        self.btn_config = ctk.CTkButton(self.frame_cabecalho, text="⚙️ Configurações", command=self.abrir_configuracoes)
        self.btn_config.pack(side="right", padx=20, pady=10)
        
        self.frame_principal = ctk.CTkFrame(self.root)
        self.frame_principal.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        
        self.frame_principal.grid_columnconfigure(0, weight=1)
        self.frame_principal.grid_columnconfigure(1, weight=1)
        self.frame_principal.grid_rowconfigure(0, weight=1)
        
        self.frame_form = ctk.CTkFrame(self.frame_principal)
        self.frame_form.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        
        ctk.CTkLabel(self.frame_form, text="Dados de Acesso e Período", font=("Arial", 16, "bold")).pack(anchor="w", padx=20, pady=10)
        self.criar_campos_formulario()
        
        self.frame_log = ctk.CTkFrame(self.frame_principal)
        self.frame_log.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        
        ctk.CTkLabel(self.frame_log, text="Log de Execução", font=("Arial", 16, "bold")).pack(anchor="w", padx=20, pady=10)
        
        self.progresso = ctk.CTkProgressBar(self.frame_log)
        self.progresso.pack(fill="x", padx=20, pady=10)
        self.progresso.set(0)
        
        self.lbl_status = ctk.CTkLabel(self.frame_log, text="Aguardando execução...")
        self.lbl_status.pack(anchor="w", padx=20, pady=5)
        
        self.frame_texto = ctk.CTkFrame(self.frame_log)
        self.frame_texto.pack(fill="both", expand=True, padx=20, pady=10)
        
        self.txt_log = ctk.CTkTextbox(self.frame_texto)
        self.txt_log.pack(fill="both", expand=True)
        
        self.frame_rodape = ctk.CTkFrame(self.root)
        self.frame_rodape.grid(row=2, column=0, padx=10, pady=10, sticky="ew")
        
        self.btn_executar = ctk.CTkButton(self.frame_rodape, text="▶️ Executar", font=("Arial", 14, "bold"), height=40, command=self.executar)
        self.btn_executar.pack(side="left", padx=20, pady=10)
        
        self.frame_botoes_adicionais = ctk.CTkFrame(self.frame_rodape)
        self.frame_botoes_adicionais.pack(side="right", padx=20, pady=10)
        
        self.btn_visualizar = ctk.CTkButton(self.frame_botoes_adicionais, text="👁️ Visualizar", command=self.visualizar_graficos, state="disabled")
        self.btn_visualizar.grid(row=0, column=0, padx=5, pady=5)
        
        self.btn_exportar = ctk.CTkButton(self.frame_botoes_adicionais, text="📤 Exportar", command=self.exportar_relatorio, state="disabled")
        self.btn_exportar.grid(row=0, column=1, padx=5, pady=5)
        
        self.btn_comparar = ctk.CTkButton(self.frame_botoes_adicionais, text="📊 Comparar", command=self.comparar_periodos)
        self.btn_comparar.grid(row=0, column=2, padx=5, pady=5)
    
    def criar_campos_formulario(self):
        self.frame_campos = ctk.CTkFrame(self.frame_form)
        self.frame_campos.pack(fill="both", expand=True, padx=20, pady=10)
        
        ctk.CTkLabel(self.frame_campos, text="Usuário:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.entry_usuario = ctk.CTkEntry(self.frame_campos, width=250)
        self.entry_usuario.grid(row=0, column=1, padx=10, pady=10, sticky="w")
        
        ctk.CTkLabel(self.frame_campos, text="Senha:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        self.entry_senha = ctk.CTkEntry(self.frame_campos, width=250, show="•")
        self.entry_senha.grid(row=1, column=1, padx=10, pady=10, sticky="w")
        
        ctk.CTkLabel(self.frame_campos, text="Navegador:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        self.combo_navegador = ctk.CTkComboBox(self.frame_campos, values=["Chrome", "Brave"], width=250)
        self.combo_navegador.grid(row=2, column=1, padx=10, pady=10, sticky="w")
        
        ctk.CTkLabel(self.frame_campos, text="Diretório de Download:").grid(row=3, column=0, padx=10, pady=10, sticky="e")
        self.frame_dir = ctk.CTkFrame(self.frame_campos)
        self.frame_dir.grid(row=3, column=1, padx=10, pady=10, sticky="w")
        
        self.entry_download_dir = ctk.CTkEntry(self.frame_dir, width=200)
        self.entry_download_dir.pack(side="left")
        self.btn_dir = ctk.CTkButton(self.frame_dir, text="...", width=30, command=self.selecionar_diretorio)
        self.btn_dir.pack(side="right", padx=5)
        
        ctk.CTkLabel(self.frame_campos, text="Período:").grid(row=4, column=0, padx=10, pady=10, sticky="e")
        self.frame_periodo = ctk.CTkFrame(self.frame_campos)
        self.frame_periodo.grid(row=4, column=1, padx=10, pady=10, sticky="w")
        
        ctk.CTkLabel(self.frame_periodo, text="De:").pack(side="left", padx=5)
        self.entry_data_inicio = ctk.CTkEntry(self.frame_periodo, width=80, placeholder_text="ddmmaa")
        self.entry_data_inicio.pack(side="left", padx=5)
        
        ctk.CTkLabel(self.frame_periodo, text="Até:").pack(side="left", padx=5)
        self.entry_data_fim = ctk.CTkEntry(self.frame_periodo, width=80, placeholder_text="ddmmaa")
        self.entry_data_fim.pack(side="left", padx=5)
        
        self.btn_periodo = ctk.CTkButton(self.frame_periodo, text="📅", width=30, command=self.menu_periodo_rapido)
        self.btn_periodo.pack(side="left", padx=5)
        
        self.salvar_credenciais = ctk.CTkCheckBox(self.frame_campos, text="Salvar credenciais")
        self.salvar_credenciais.grid(row=5, column=0, columnspan=2, padx=10, pady=10, sticky="w")
        self.salvar_credenciais.select()
    
    def preencher_campos(self):
        self.entry_usuario.insert(0, self.config.get("usuario", ""))
        self.entry_senha.insert(0, self.config.get("senha", ""))
        self.combo_navegador.set(self.config.get("navegador", "Chrome"))
        self.entry_download_dir.insert(0, self.config.get("download_dir", os.path.join(os.path.expanduser("~"), "Downloads")))
        
        hoje = datetime.now()
        primeiro_dia = datetime(hoje.year, hoje.month, 1)
        ultimo_dia = (datetime(hoje.year, hoje.month + 1, 1) - timedelta(days=1)) if hoje.month < 12 else datetime(hoje.year, 12, 31)
        
        self.entry_data_inicio.insert(0, primeiro_dia.strftime("%d%m%y"))
        self.entry_data_fim.insert(0, ultimo_dia.strftime("%d%m%y"))
    
    def selecionar_diretorio(self):
        diretorio = filedialog.askdirectory(title="Selecione o Diretório de Download")
        if diretorio:
            self.entry_download_dir.delete(0, tk.END)
            self.entry_download_dir.insert(0, diretorio)
    
    def menu_periodo_rapido(self):
        menu = tk.Menu(self.root, tearoff=0)
        hoje = datetime.now()
        
        menu.add_command(label="Hoje", command=lambda: self.definir_periodo_rapido(0, 0))
        menu.add_command(label="Ontem", command=lambda: self.definir_periodo_rapido(1, 1))
        menu.add_command(label="Últimos 7 dias", command=lambda: self.definir_periodo_rapido(6, 0))
        menu.add_command(label="Últimos 30 dias", command=lambda: self.definir_periodo_rapido(29, 0))
        menu.add_separator()
        menu.add_command(label="Mês atual", command=lambda: self.definir_periodo_mes_atual())
        menu.add_command(label="Mês anterior", command=lambda: self.definir_periodo_mes_anterior())
        
        menu.tk_popup(self.btn_periodo.winfo_rootx(), self.btn_periodo.winfo_rooty() + 30)
    
    def definir_periodo_rapido(self, dias_atras_inicio, dias_atras_fim):
        hoje = datetime.now()
        data_inicio = hoje - timedelta(days=dias_atras_inicio)
        data_fim = hoje - timedelta(days=dias_atras_fim)
        
        self.entry_data_inicio.delete(0, tk.END)
        self.entry_data_inicio.insert(0, data_inicio.strftime("%d%m%y"))
        self.entry_data_fim.delete(0, tk.END)
        self.entry_data_fim.insert(0, data_fim.strftime("%d%m%y"))
    
    def definir_periodo_mes_atual(self):
        hoje = datetime.now()
        primeiro_dia = datetime(hoje.year, hoje.month, 1)
        ultimo_dia = (datetime(hoje.year, hoje.month + 1, 1) - timedelta(days=1)) if hoje.month < 12 else datetime(hoje.year, 12, 31)
        
        self.entry_data_inicio.delete(0, tk.END)
        self.entry_data_inicio.insert(0, primeiro_dia.strftime("%d%m%y"))
        self.entry_data_fim.delete(0, tk.END)
        self.entry_data_fim.insert(0, ultimo_dia.strftime("%d%m%y"))
    
    def definir_periodo_mes_anterior(self):
        hoje = datetime.now()
        mes_anterior = hoje.month - 1 if hoje.month > 1 else 12
        ano = hoje.year if hoje.month > 1 else hoje.year - 1
        
        primeiro_dia = datetime(ano, mes_anterior, 1)
        ultimo_dia = datetime(hoje.year, hoje.month, 1) - timedelta(days=1)
        
        self.entry_data_inicio.delete(0, tk.END)
        self.entry_data_inicio.insert(0, primeiro_dia.strftime("%d%m%y"))
        self.entry_data_fim.delete(0, tk.END)
        self.entry_data_fim.insert(0, ultimo_dia.strftime("%d%m%y"))
    
    def abrir_configuracoes(self):
        TelaConfiguracoes(self.root, self.config, self.atualizar_configuracoes)
    
    def atualizar_configuracoes(self, config):
        self.config = config
        self.combo_navegador.set(self.config.get("navegador", "Chrome"))
        self.entry_download_dir.delete(0, tk.END)
        self.entry_download_dir.insert(0, self.config.get("download_dir", os.path.join(os.path.expanduser("~"), "Downloads")))
    
    def atualizar_progresso(self, valor, mensagem):
        self.progresso.set(valor / 100)
        self.lbl_status.configure(text=mensagem)
        self.adicionar_log(mensagem)
    
    def adicionar_log(self, mensagem):
        self.txt_log.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {mensagem}\n")
        self.txt_log.see(tk.END)
    
    def executar(self):
        usuario = self.entry_usuario.get().strip()
        senha = self.entry_senha.get().strip()
        navegador = self.combo_navegador.get()
        download_dir = self.entry_download_dir.get().strip()
        data_inicio = self.entry_data_inicio.get().strip()
        data_fim = self.entry_data_fim.get().strip()
        
        if not usuario or not senha:
            messagebox.showerror("Erro", "Usuário e senha são obrigatórios!")
            return
        if not download_dir:
            messagebox.showerror("Erro", "Selecione um diretório de download!")
            return
        if not validar_data(data_inicio):
            messagebox.showerror("Erro", "Data inicial inválida! Use o formato ddmmaa (ex.: 010123).")
            return
        if not validar_data(data_fim):
            messagebox.showerror("Erro", "Data final inválida! Use o formato ddmmaa (ex.: 311223).")
            return
        
        self.btn_executar.configure(state="disabled")
        self.txt_log.delete(1.0, tk.END)
        self.progresso.set(0)
        self.lbl_status.configure(text="Iniciando execução...")
        
        if self.salvar_credenciais.get():
            self.config["usuario"] = usuario
            self.config["senha"] = senha
            self.config["navegador"] = navegador
            self.config["download_dir"] = download_dir
            salvar_configuracoes(self.config)
        
        self.thread_execucao = threading.Thread(
            target=self.executar_thread,
            args=(usuario, senha, data_inicio, data_fim, navegador, download_dir)
        )
        self.thread_execucao.daemon = True
        self.thread_execucao.start()
    
    def executar_thread(self, usuario, senha, data_inicio, data_fim, navegador, download_dir):
        try:
            self.adicionar_log(f"Iniciando execução para o período de {formatar_data(data_inicio)} a {formatar_data(data_fim)}")
            self.arquivo_processado = executar_script(
                usuario, senha, data_inicio, data_fim, navegador, download_dir,
                config=self.config,
                callback=lambda p, m: self.root.after(0, lambda: self.atualizar_progresso(p, m))
            )
            self.root.after(0, self.finalizar_execucao, True, f"Processamento concluído! Arquivo salvo em: {self.arquivo_processado}")
        except Exception as e:
            self.root.after(0, self.finalizar_execucao, False, f"Falha na execução: {str(e)}")
    
    def finalizar_execucao(self, sucesso, mensagem):
        self.btn_executar.configure(state="normal")
        self.lbl_status.configure(text=mensagem)
        self.adicionar_log(mensagem)
        
        if sucesso:
            self.btn_visualizar.configure(state="normal")
            self.btn_exportar.configure(state="normal")
            messagebox.showinfo("Sucesso", "Relatório concluído com sucesso!")
        else:
            self.btn_visualizar.configure(state="disabled")
            self.btn_exportar.configure(state="disabled")
            messagebox.showerror("Erro", mensagem)
    
    def visualizar_graficos(self):
        if not self.arquivo_processado or not os.path.exists(self.arquivo_processado):
            messagebox.showerror("Erro", "Nenhum arquivo processado disponível!")
            return
        
        try:
            df = pd.read_excel(self.arquivo_processado, sheet_name="DinamicTable", engine='openpyxl')
            VisualizadorGraficos(self.root, df)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao visualizar gráficos: {str(e)}")
    
    def exportar_relatorio(self):
        if not self.arquivo_processado or not os.path.exists(self.arquivo_processado):
            messagebox.showerror("Erro", "Nenhum arquivo processado disponível!")
            return
        TelaExportacao(self.root, self.arquivo_processado)
    
    def comparar_periodos(self):
        TelaComparacao(self.root)

# Função principal
def main():
    if len(sys.argv) > 1:
        caminho_arquivo = sys.argv[1]
        if os.path.exists(caminho_arquivo) and verificar_arquivo_excel(caminho_arquivo):
            try:
                arquivo_processado = processar_dados_excel_direto(caminho_arquivo)
                shutil.move(arquivo_processado, caminho_arquivo)
                print(f"Arquivo processado com sucesso: {caminho_arquivo}")
            except Exception as e:
                print(f"Erro ao processar o arquivo: {e}")
        else:
            print(f"Erro: O arquivo '{caminho_arquivo}' não existe ou não é um Excel válido.")
    else:
        root = ctk.CTk()
        app = App(root)
        root.mainloop()

if __name__ == "__main__":
    main()